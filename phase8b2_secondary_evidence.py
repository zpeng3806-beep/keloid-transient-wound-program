#!/usr/bin/env python3
"""Create prespecified Phase 8B2 sensitivity and secondary evidence tables."""

from __future__ import annotations

import argparse
import csv
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


PRIMARY_TISSUES = {
    "Cells_Cultured_fibroblasts",
    "Skin_Sun_Exposed_Lower_leg",
    "Skin_Not_Sun_Exposed_Suprapubic",
}


def read_signature(path: Path) -> list[str]:
    with path.open(newline="") as src:
        return [row["gene"].strip() for row in csv.DictReader(src, delimiter="\t")]


def read_mapping(path: Path) -> list[dict[str, str]]:
    with path.open(newline="") as src:
        return list(csv.DictReader(src, delimiter="\t"))


def read_gene_p(path: Path) -> dict[str, str]:
    values: dict[str, str] = {}
    with path.open() as src:
        header = src.readline().split()
        gene_i, p_i = header.index("GENE"), header.index("P")
        for line in src:
            row = line.split()
            values[row[gene_i]] = row[p_i]
    return values


def read_gsa(path: Path) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    with path.open() as src:
        header = None
        for line in src:
            if line.startswith("VARIABLE"):
                header = line.split()
            elif header and line.strip() and not line.startswith("#"):
                rows.append(dict(zip(header, line.split())))
    return rows


def bh_adjust(ps: list[float]) -> list[float]:
    n = len(ps)
    order = sorted(range(n), key=lambda i: ps[i])
    adjusted = [1.0] * n
    running = 1.0
    for rank_index in range(n - 1, -1, -1):
        i = order[rank_index]
        rank = rank_index + 1
        running = min(running, ps[i] * n / rank)
        adjusted[i] = min(1.0, running)
    return adjusted


def rows_from_sheet(path: Path, sheet_name: str) -> list[dict[str, object]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    headers = [cell.value for cell in ws[2]]
    rows: list[dict[str, object]] = []
    for values in ws.iter_rows(min_row=3, values_only=True):
        if not any(value is not None for value in values):
            continue
        rows.append({str(headers[i]): value for i, value in enumerate(values) if i < len(headers) and headers[i] is not None})
    wb.close()
    return rows


def write_tsv(path: Path, fields: list[str], rows: list[dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="") as out:
        writer = csv.DictWriter(out, fieldnames=fields, delimiter="\t", lineterminator="\n", extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--project", type=Path, required=True)
    args = parser.parse_args()
    p = args.project
    results = p / "06_results/genetics"
    supplements = p / "data/external/genetics/Greene2025_supplementary"

    core = read_signature(p / "06_results/gateA/TWPS_PRIMARY_D7_M3_CORE100.tsv")
    core_set = set(core)
    mapping = read_mapping(results / "CORE100_MAGMA_MAPPING.tsv")
    gene_id_by_symbol = {row["locked_gene"]: row["MAGMA_gene_id"] for row in mapping if row["MAGMA_gene_id"]}
    eur_p = read_gene_p(results / "magma/EUR_gene_results.genes.out")
    afr_p = read_gene_p(results / "magma/AFR_gene_results.genes.out")

    sensitivity_rows: list[dict[str, object]] = []
    for ancestry in ("EUR", "AFR"):
        for row in read_gsa(results / f"magma/{ancestry}_signature_sensitivity.gsa.out"):
            beta = float(row["BETA"])
            sensitivity_rows.append(
                {
                    "signature": row["VARIABLE"],
                    "ancestry": ancestry,
                    "mapped_n": int(row["NGENES"]),
                    "beta": beta,
                    "SE": float(row["SE"]),
                    "P": float(row["P"]),
                    "direction": "POSITIVE" if beta > 0 else "NEGATIVE" if beta < 0 else "ZERO",
                }
            )
    fdrs = bh_adjust([float(row["P"]) for row in sensitivity_rows])
    for row, fdr in zip(sensitivity_rows, fdrs):
        row["BH_FDR"] = fdr
    write_tsv(results / "MAGMA_SIGNATURE_SENSITIVITY.tsv", ["signature", "ancestry", "mapped_n", "beta", "SE", "P", "direction", "BH_FDR"], sensitivity_rows)

    # Data 6b exposes only genes occurring in reported significant GENE2FUNC terms.
    # This is a verifiable 115-gene subset, not the full author-stated 119-gene mapping list.
    data6 = supplements / "41467_2025_62945_MOESM8_ESM.xlsx"
    fuma_rows = rows_from_sheet(data6, "SD6b. FUMA GENE2FUNC Multi")
    recoverable_fuma = set()
    for row in fuma_rows:
        genes = row.get("genes")
        if genes:
            recoverable_fuma.update(g.strip() for g in str(genes).split(":") if g.strip())
    recoverable_rows = [{"gene": gene, "CORE100_overlap": "YES" if gene in core_set else "NO"} for gene in sorted(recoverable_fuma)]
    write_tsv(results / "FUMA_GENE2FUNC_MULTI_RECOVERABLE_115.tsv", ["gene", "CORE100_overlap"], recoverable_rows)

    data7 = supplements / "41467_2025_62945_MOESM9_ESM.xlsx"
    all_gpge = rows_from_sheet(data7, "SD7. All sig GPGE results")
    gpge_rows: list[dict[str, object]] = []
    for row in all_gpge:
        gene = str(row.get("gene_name") or "")
        tissue = str(row.get("tissue") or "")
        if gene not in core_set or tissue not in PRIMARY_TISSUES:
            continue
        effect = float(row["effect_size"])
        gpge_rows.append(
            {
                "gene": gene,
                "tissue": tissue,
                "ancestry_analysis": row.get("analysis"),
                "effect_direction": "INCREASED_RISK" if effect > 0 else "DECREASED_RISK" if effect < 0 else "ZERO",
                "effect_size": effect,
                "P": row.get("pvalue"),
                "reported_significance": "AUTHOR_REPORTED_SIGNIFICANT",
                "TWPS_directional_class": "CONCORDANT_WITH_TWPS" if effect > 0 else "DISCORDANT_WITH_TWPS" if effect < 0 else "UNRESOLVED",
            }
        )
    write_tsv(results / "CORE100_GPGE_CONVERGENCE.tsv", ["gene", "tissue", "ancestry_analysis", "effect_direction", "effect_size", "P", "reported_significance", "TWPS_directional_class"], gpge_rows)

    significant_gpge_pairs = {(str(r["gene"]), str(r["tissue"]), str(r["ancestry_analysis"])) for r in gpge_rows}
    data8 = supplements / "41467_2025_62945_MOESM10_ESM.xlsx"
    sheet_analysis = {
        "SD8a. TWAS coloc Multi": "multi",
        "SD8b. TWAS coloc EUR": "EUR",
        "SD8c. TWAS coloc EAS": "EAS",
        "SD8d. TWAS coloc AFR": "AFR",
    }
    coloc_rows: list[dict[str, object]] = []
    for sheet, analysis in sheet_analysis.items():
        for row in rows_from_sheet(data8, sheet):
            gene = str(row.get("GeneName") or "")
            tissue = str(row.get("Tissue") or "")
            pair = (gene, tissue, analysis)
            pp = row.get("PP.H4.abf")
            if pair not in significant_gpge_pairs or pp is None or float(pp) <= 0.90:
                continue
            effect = float(row["effect_size"])
            coloc_rows.append(
                {
                    "gene": gene,
                    "tissue": tissue,
                    "ancestry_analysis": analysis,
                    "effect_direction": "INCREASED_RISK" if effect > 0 else "DECREASED_RISK" if effect < 0 else "ZERO",
                    "GPGE_P": row.get("p-value"),
                    "posterior_probability": float(pp),
                    "classification": "HIGHER_CONFIDENCE_GENETIC_EXPRESSION_CONVERGENCE",
                }
            )
    write_tsv(results / "CORE100_COLOCALIZATION.tsv", ["gene", "tissue", "ancestry_analysis", "effect_direction", "GPGE_P", "posterior_probability", "classification"], coloc_rows)

    # Data 3 and Data 2 are variant/locus tables without a complete author-supported
    # mapping from every row to the CORE100 genes. Preserve the boundary explicitly.
    boundary_rows = [{"gene": gene, "support": "NOT_ASSIGNABLE", "notes": "No exact author-supported gene-locus relationship in the local supplementary table; nearest-gene assignment prohibited"} for gene in core]
    write_tsv(results / "CORE100_FINEMAPPING_SUPPORT.tsv", ["gene", "support", "notes"], boundary_rows)
    write_tsv(results / "CORE100_ALLOFUS_REPLICATION_SUPPORT.tsv", ["gene", "support", "notes"], boundary_rows)

    gpge_by_gene: dict[str, list[dict[str, object]]] = defaultdict(list)
    coloc_by_gene: dict[str, list[dict[str, object]]] = defaultdict(list)
    for row in gpge_rows:
        gpge_by_gene[str(row["gene"])].append(row)
    for row in coloc_rows:
        coloc_by_gene[str(row["gene"])].append(row)

    matrix_rows: list[dict[str, object]] = []
    for gene in core:
        gene_id = gene_id_by_symbol.get(gene, "")
        gpg = gpge_by_gene.get(gene, [])
        col = coloc_by_gene.get(gene, [])
        if gene in recoverable_fuma:
            fuma_state = "YES_IN_RECOVERABLE_115_SUBSET"
        else:
            fuma_state = "UNRESOLVED_FULL_119_LIST_UNAVAILABLE"
        matrix_rows.append(
            {
                "gene": gene,
                "EUR_MAGMA_gene_P": eur_p.get(gene_id, ""),
                "AFR_MAGMA_gene_P": afr_p.get(gene_id, ""),
                "FUMA119": fuma_state,
                "GPGE_relevant_tissue": ";".join(str(r["tissue"]) for r in gpg),
                "GPGE_direction": ";".join(str(r["effect_direction"]) for r in gpg),
                "GPGE_P": ";".join(str(r["P"]) for r in gpg),
                "colocalization_PP": ";".join(str(r["posterior_probability"]) for r in col),
                "strong_colocalization": "YES" if col else "NO",
                "fine_mapping_support": "NOT_ASSIGNABLE",
                "AllOfUs_replication_support": "NOT_ASSIGNABLE",
                "evidence_notes": "Absence from Data7 means not reported as significant, not no genetic effect; exact FUMA119 list unavailable locally",
            }
        )
    write_tsv(results / "CORE100_GENETIC_EVIDENCE_MATRIX.tsv", ["gene", "EUR_MAGMA_gene_P", "AFR_MAGMA_gene_P", "FUMA119", "GPGE_relevant_tissue", "GPGE_direction", "GPGE_P", "colocalization_PP", "strong_colocalization", "fine_mapping_support", "AllOfUs_replication_support", "evidence_notes"], matrix_rows)


if __name__ == "__main__":
    main()
